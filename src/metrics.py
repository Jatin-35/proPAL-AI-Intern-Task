import openpyxl
import os
from datetime import datetime

class MetricsLogger:
    def __init__(self, filename="metrics.xlsx"):
        self.filename = filename
        if not os.path.exists(self.filename):
            self._create_file()

    def _create_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Latency Logs"
        ws.append([
            "Timestamp", "Session ID", "Speech ID",
            "EOU Delay (s)", "TTFT (s)", "TTFB (s)", "Total Latency (s)"
        ])
        wb.save(self.filename)

    def log_latency(self, session_id, speech_id, eou_delay, ttft, ttfb):
        total_latency = (
            (eou_delay or 0.0) + (ttft or 0.0) + (ttfb or 0.0)
        )
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            session_id,
            speech_id,
            eou_delay,
            ttft,
            ttfb,
            total_latency
        ])
        wb.save(self.filename)

    def log_usage_summary(self, summary: dict):
        wb = openpyxl.load_workbook(self.filename)
        if "Summary" not in wb.sheetnames:
            ws = wb.create_sheet("Summary")
            ws.append(["Key", "Value"])
        else:
            ws = wb["Summary"]
        for k, v in summary.items():
            ws.append([k, v])
        wb.save(self.filename)
